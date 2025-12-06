# backend/app/core/document_parser.py
"""Document parsing for PDF, CSV, Excel, and text files."""

from __future__ import annotations

import csv
import io
from typing import Optional

from app.core.observability import logs


class DocumentParser:
    """Parse documents to extract text."""

    SUPPORTED_TYPES = {
        "application/pdf": "pdf",
        "text/plain": "text",
        "text/markdown": "text",
        "text/csv": "csv",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "application/vnd.ms-excel": "xls",
        "application/json": "text",
        "application/octet-stream": "auto",
    }

    async def parse(self, content: bytes, content_type: str, filename: str) -> str:
        """Parse document content to text."""
        try:
            file_type = self.SUPPORTED_TYPES.get(content_type, "auto")

            # Auto-detect from extension
            if file_type == "auto":
                lower_filename = filename.lower()
                if lower_filename.endswith(".pdf"):
                    file_type = "pdf"
                elif lower_filename.endswith(".csv"):
                    file_type = "csv"
                elif lower_filename.endswith(".xlsx"):
                    file_type = "xlsx"
                elif lower_filename.endswith(".xls"):
                    file_type = "xls"
                else:
                    file_type = "text"

            if file_type == "pdf":
                return await self._parse_pdf(content, filename)
            elif file_type == "csv":
                return self._parse_csv(content, filename)
            elif file_type in ("xlsx", "xls"):
                return await self._parse_excel(content, filename)
            else:
                return self._parse_text(content, filename)

        except Exception as e:
            logs.error(f"Failed to parse document: {filename}", "parser", exception=e)
            # Fallback to raw text
            return content.decode("utf-8", errors="ignore")

    async def _parse_pdf(self, content: bytes, filename: str) -> str:
        """Extract text from PDF using pypdf."""
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(content))
        text_parts = []

        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)

        text = "\n\n".join(text_parts)

        logs.info(
            f"Parsed PDF: {filename}",
            "parser",
            metadata={"chars": len(text), "pages": len(reader.pages)},
        )

        return text

    def _parse_text(self, content: bytes, filename: str) -> str:
        """Parse text/markdown file."""
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1", errors="ignore")

        logs.info(
            f"Parsed text: {filename}",
            "parser",
            metadata={"chars": len(text)},
        )

        return text

    def _parse_csv(self, content: bytes, filename: str) -> str:
        """Parse CSV file to text representation with table structure preserved."""
        try:
            text_content = content.decode("utf-8")
        except UnicodeDecodeError:
            text_content = content.decode("latin-1", errors="ignore")

        # Parse CSV and convert to readable text
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)

        if not rows:
            return ""

        # Build a text representation that preserves table structure
        text_parts = []
        headers = rows[0] if rows else []

        # Add headers
        if headers:
            text_parts.append("Table headers: " + " | ".join(headers))
            text_parts.append("")

        # Add rows with labeled columns
        for i, row in enumerate(rows[1:], 1):
            row_parts = []
            for j, cell in enumerate(row):
                header = headers[j] if j < len(headers) else f"Column {j+1}"
                row_parts.append(f"{header}: {cell}")
            text_parts.append(f"Row {i}: " + " | ".join(row_parts))

        text = "\n".join(text_parts)

        logs.info(
            f"Parsed CSV: {filename}",
            "parser",
            metadata={"chars": len(text), "rows": len(rows)},
        )

        return text

    async def _parse_excel(self, content: bytes, filename: str) -> str:
        """Parse Excel file to text representation."""
        try:
            import openpyxl
        except ImportError:
            logs.warning("openpyxl not installed, falling back to raw text", "parser")
            return self._parse_text(content, filename)

        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")

            # Get all rows
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = rows[0] if rows else []

            # Add headers
            if headers:
                header_strs = [str(h) if h is not None else "" for h in headers]
                text_parts.append("Table headers: " + " | ".join(header_strs))
                text_parts.append("")

            # Add rows with labeled columns
            for i, row in enumerate(rows[1:], 1):
                row_parts = []
                for j, cell in enumerate(row):
                    header = headers[j] if j < len(headers) and headers[j] else f"Column {j+1}"
                    cell_val = str(cell) if cell is not None else ""
                    row_parts.append(f"{header}: {cell_val}")
                text_parts.append(f"Row {i}: " + " | ".join(row_parts))

        text = "\n".join(text_parts)

        logs.info(
            f"Parsed Excel: {filename}",
            "parser",
            metadata={"chars": len(text), "sheets": len(workbook.sheetnames)},
        )

        return text


_parser: Optional[DocumentParser] = None


def get_parser() -> DocumentParser:
    """Get document parser instance."""
    global _parser
    if _parser is None:
        _parser = DocumentParser()
    return _parser
